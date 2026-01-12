import os
from io import BytesIO
from datetime import date, timedelta

import pandas as pd
import psycopg2
import streamlit as st
import plotly.express as px
import base64

# -----------------------------
# Config & UI Aesthetics
# -----------------------------
st.set_page_config(
    page_title="EXPERA | Inteligencia de Negocios",
    page_icon="💎",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# Custom CSS for an Ultra-Premium Look
st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');
    
    :root {
        --primary: #3b82f6;
        --secondary: #6366f1;
        --accent: #8b5cf6;
        --bg-light: #f8fafc;
        --text-main: #0f172a;
        --text-sub: #64748b;
        --card-bg: #ffffff;
    }

    /* Global Styles & Light Mode Enforcement */
    .stApp {
        background: radial-gradient(circle at top right, #f1f5f9, #ffffff) !important;
        color: #1e293b !important;
    }
    
    html, body, [data-testid="stSidebar"], [data-testid="stAppViewContainer"] {
        font-family: 'Inter', sans-serif;
        background-color: white !important;
    }

    /* Force visibility of all text labels and headers */
    label, p, span, h1, h2, h3, h4, .stMarkdown {
        color: #1e293b !important;
    }

    /* Clean Logo Container */
    .main-header {
        background-color: transparent;
        padding: 10px 20px 20px 20px;
        text-align: center;
        margin-bottom: 10px;
    }

    /* Move app content higher */
    .block-container {
        padding-top: 1rem !important;
    }

    /* Hide default elements */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}
    
    /* Metrics Dashboard Positioning */
    div[data-testid="stMetric"] {
        background: rgba(255, 255, 255, 0.8) !important;
        backdrop-filter: blur(10px) !important;
        padding: 24px !important;
        border-radius: 20px !important;
        border: 1px solid rgba(226, 232, 240, 0.5) !important;
        box-shadow: 0 10px 15px -3px rgba(0, 0, 0, 0.05) !important;
        transition: all 0.4s cubic-bezier(0.4, 0, 0.2, 1) !important;
    }
    
    div[data-testid="stMetric"]:hover {
        transform: translateY(-5px);
        box-shadow: 0 20px 25px -5px rgba(59, 130, 246, 0.1) !important;
        border-color: var(--primary) !important;
    }

    [data-testid="stMetricValue"] {
        font-weight: 800;
        font-size: 2.2rem !important;
        color: var(--text-main) !important;
        letter-spacing: -0.02em;
    }
    
    [data-testid="stMetricLabel"] {
        font-weight: 600;
        text-transform: uppercase;
        letter-spacing: 0.1em;
        font-size: 0.7rem !important;
        color: var(--text-sub) !important;
    }
    
    /* Unified Search Bar & Filter Button Style */
    div[data-testid="stTextInput"] > div[data-baseweb="input"],
    div[data-testid="stPopover"] > button,
    .stTextInput input {
        border-radius: 20px !important;
        border: 1px solid rgba(226, 232, 240, 0.8) !important;
        background-color: white !important;
        color: #1e293b !important;
        box-shadow: 0 10px 30px -10px rgba(0,0,0,0.05) !important;
        transition: all 0.4s cubic-bezier(0.4, 0, 0.2, 1) !important;
    }
    
    div[data-testid="stTextInput"] input {
        padding: 0 30px !important;
        font-size: 1.1rem !important;
        height: 60px !important;
        background-color: white !important;
        color: #1e293b !important;
    }

    div[data-testid="stPopover"] > button {
        width: 100% !important;
        color: var(--text-sub) !important;
        font-weight: 500 !important;
        font-size: 1rem !important;
        padding: 0 25px !important;
    }
    
    /* Interactive States (Unified) */
    div[data-testid="stTextInput"] > div[data-baseweb="input"]:focus-within,
    div[data-testid="stPopover"] > button:hover {
        border-color: var(--primary) !important;
        box-shadow: 0 0 0 4px rgba(59, 130, 246, 0.1) !important;
        transform: translateY(-2px);
        color: var(--primary) !important;
    }

    /* Feature Cards (Landing Page) */
    .feature-card {
        background: rgba(255, 255, 255, 0.7) !important;
        backdrop-filter: blur(12px) !important;
        padding: 48px 32px;
        border-radius: 32px;
        border: 1px solid rgba(255, 255, 255, 0.4);
        text-align: center;
        height: 340px;
        box-shadow: 0 10px 25px -5px rgba(0, 0, 0, 0.05);
        transition: all 0.5s cubic-bezier(0.4, 0, 0.2, 1);
        display: flex;
        flex-direction: column;
        justify-content: center;
        align-items: center;
    }

    .feature-card:hover {
        transform: translateY(-12px) scale(1.02);
        background: white !important;
        box-shadow: 0 30px 50px -12px rgba(59, 130, 246, 0.15);
        border-color: var(--primary);
    }

    .feature-card-icon {
        font-size: 54px;
        margin-bottom: 24px;
        filter: drop-shadow(0 10px 10px rgba(0,0,0,0.1));
    }

    .feature-card h3 {
        color: var(--text-main);
        margin-bottom: 12px;
        font-weight: 800;
        font-size: 1.5rem;
    }

    .feature-card p {
        color: var(--text-sub);
        line-height: 1.7;
        font-size: 0.95rem;
    }

    /* Login Box */
    .login-container {
        max-width: 450px;
        margin: 100px auto;
        padding: 40px;
        background: white;
        border-radius: 32px;
        box-shadow: 0 25px 50px -12px rgba(0, 0, 0, 0.1);
        text-align: center;
        border: 1px solid #e2e8f0;
    }

    .stForm {
        background-color: white !important;
        border-radius: 24px !important;
        border: 1px solid #f1f5f9 !important;
    }

    /* Interactive Popover Style */
    div[data-testid="stPopover"] > button {
        height: 60px !important;
        min-height: 60px !important;
        width: 100% !important;
        color: #64748b !important;
        font-weight: 500 !important;
        font-size: 1rem !important;
        padding: 0 25px !important;
        background-color: white !important;
        border: 2px solid #f1f5f9 !important;
    }

    /* Interactive Buttons */
    .stButton button {
        border-radius: 16px !important;
        height: 55px !important;
        font-weight: 700 !important;
        background: linear-gradient(135deg, var(--primary) 0%, var(--secondary) 100%) !important;
        color: white !important;
        border: none !important;
        padding: 0 30px !important;
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1) !important;
        text-transform: uppercase;
        letter-spacing: 1px;
    }
    
    .stButton button:hover {
        transform: translateY(-2px);
        box-shadow: 0 10px 20px rgba(59, 130, 246, 0.3) !important;
        filter: brightness(1.1);
    }

    /* Modern Tabs */
    .stTabs [data-baseweb="tab-list"] {
        background-color: #f1f5f9;
        border-radius: 16px;
        padding: 6px;
        gap: 10px;
    }

    .stTabs [data-baseweb="tab"] {
        height: 45px;
        background-color: transparent;
        border-radius: 12px;
        font-weight: 700;
        color: var(--text-sub);
        border: none !important;
        padding: 0 20px !important;
    }

    .stTabs [aria-selected="true"] {
        background-color: white !important;
        color: var(--primary) !important;
        box-shadow: 0 4px 12px rgba(0, 0, 0, 0.05) !important;
    }

    /* Custom Table Style */
    .stDataFrame {
        border: 1px solid #e2e8f0;
        border-radius: 20px;
        overflow: hidden;
    }

    /* Card Containers */
    .info-card {
        background: white;
        padding: 25px;
        border-radius: 20px;
        border: 1px solid #ecf0f1;
        box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.02);
        margin-bottom: 20px;
    }

    .info-card h3 {
        color: var(--text-main);
        margin-top: 0;
        font-size: 1.25rem;
    }
    
    /* Animation for results */
    @keyframes fadeIn {
        from { opacity: 0; transform: translateY(10px); }
        to { opacity: 1; transform: translateY(0); }
    }
    .fade-in {
        animation: fadeIn 0.5s ease-out forwards;
    }
    </style>
""", unsafe_allow_html=True)

def get_env(name: str, default: str | None = None) -> str | None:
    v = os.getenv(name, default)
    if v is None or str(v).strip() == "":
        return default
    return str(v)

# DB Credentials
PG_HOST = get_env("PG_HOST", "localhost")
PG_PORT = int(get_env("PG_PORT", "5432"))
PG_DB = get_env("PG_DB", "upgradedb")
PG_USER = get_env("PG_USER", "postgres")
PG_PASSWORD = get_env("PG_PASSWORD", "")

# -----------------------------
# Database Engine
# -----------------------------
def get_conn():
    if not PG_PASSWORD:
        st.error("🔑 Error: PG_PASSWORD no está configurada.")
        return None
    try:
        return psycopg2.connect(
            host=PG_HOST,
            port=PG_PORT,
            database=PG_DB,
            user=PG_USER,
            password=PG_PASSWORD,
            connect_timeout=5,
        )
    except Exception as e:
        st.error(f"🔌 Error de conexión: {str(e)}")
        return None

@st.cache_data(ttl=600, show_spinner=False)
def perform_search(search_query: str, years: int = 5) -> pd.DataFrame:
    """
    Optimized SQL query focused on performance and relevance.
    Fetches all matching records without limits.
    """
    conn = get_conn()
    if not conn:
        return pd.DataFrame()

    sql = """
    SELECT
      cab.creado::date          AS "Fecha",
      prod.nombre               AS "Producto",
      cli.nombre                AS "Cliente",
      det.cantidad              AS "Cant.",
      det.precio_unitario_venta AS "Precio Unit.",
      (det.cantidad * det.precio_unitario_venta) AS "Importe Total",
      cab.numero                AS "Nº Documento",
      alm.nombre                AS "Almacén",
      vc.enlace_pdf             AS "Enlace PDF"
    FROM cmrlz.notas_pedido_cab cab
    JOIN cmrlz.notas_pedido_det det ON det.nota_pedido_id = cab.id
    JOIN extcs.productos prod       ON prod.id = det.producto_id
    LEFT JOIN tcros.direcciones dir ON dir.id = cab.direccion_cliente_id
    LEFT JOIN tcros.personas cli    ON cli.id = dir.persona_id
    LEFT JOIN extcs.almacenes alm   ON alm.id = cab.almacen_id
    LEFT JOIN cmrlz.ventas_cab vc   ON vc.id = cab.venta_id
    WHERE
      cab.anulada IS FALSE
      AND cab.venta_id IS NOT NULL
      AND prod.servicio = FALSE
      AND cab.creado >= %(min_date)s
    """
    
    params = {
        "min_date": date.today() - timedelta(days=365 * years)
    }

    if search_query.strip():
        sql += " AND (prod.nombre ILIKE %(q)s OR cli.nombre ILIKE %(q)s OR cab.numero::text ILIKE %(q)s)"
        params["q"] = f"%{search_query.strip()}%"
    
    sql += " ORDER BY cab.creado DESC"

    try:
        df = pd.read_sql_query(sql, conn, params=params)
        return df
    except Exception:
        return pd.DataFrame()
    finally:
        conn.close()

def to_excel_bytes(df: pd.DataFrame) -> bytes:
    """
    Generates a professional Excel file with formatting, auto-column widths,
    and a summary total row.
    """
    output = BytesIO()
    
    # Create the Excel writer
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Reporte Detallado")
        
        workbook = writer.book
        worksheet = writer.sheets["Reporte Detallado"]
        
        # Styles
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        total_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        total_font = Font(bold=True, size=12)
        currency_format = '"S/" #,##0.00'
        
        # Format Headers
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Auto-adjust column widths and apply formats
        for i, col in enumerate(df.columns):
            column_letter = chr(65 + i) if i < 26 else f"{chr(64 + i // 26)}{chr(65 + i % 26)}"
            
            # Find max length in column
            max_length = max(
                df[col].astype(str).map(len).max(),
                len(str(col))
            ) + 4
            worksheet.column_dimensions[column_letter].width = min(max_length, 50)
            
            # Apply format to 'Fecha'
            if "Fecha" in col:
                # Ensure data is datetime for Excel to format it correctly
                # (Streamlit/Pandas might pass it as objects)
                for row in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row, column=i+1)
                    # Use standard Excel date format string
                    cell.number_format = 'dd/mm/yyyy'
            
            # Apply currency format to 'Precio Unit.' and 'Importe Total'
            if "Precio" in col or "Importe" in col:
                for row in range(2, len(df) + 2):
                    worksheet.cell(row=row, column=i+1).number_format = currency_format

        # Add Summary Row
        last_row = len(df) + 2
        
        # Calculate and place the sum in the 'Importe Total' column first
        try:
            importe_col_idx = list(df.columns).index("Importe Total") + 1
            total_sum = df["Importe Total"].sum()
            
            # Merge label from col 1 to the cell before the total
            if importe_col_idx > 1:
                worksheet.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=importe_col_idx-1)
            
            # Label
            label_cell = worksheet.cell(row=last_row, column=1, value="RESUMEN TOTAL")
            label_cell.font = total_font
            label_cell.alignment = Alignment(horizontal="right")
            
            # Value
            total_cell = worksheet.cell(row=last_row, column=importe_col_idx, value=total_sum)
            total_cell.font = total_font
            total_cell.fill = total_fill
            total_cell.number_format = currency_format
            total_cell.alignment = Alignment(horizontal="right")
        except (ValueError, KeyError):
            # Fallback if column not found
            worksheet.cell(row=last_row, column=1, value="RESUMEN TOTAL (Columna no encontrada)").font = total_font

    return output.getvalue()

# -----------------------------
# Authentication System
# -----------------------------
def check_password():
    """Returns True if the user had the correct password."""
    def password_entered():
        """Checks whether a password entered by the user is correct."""
        if (
            st.session_state["username"] == "user"
            and st.session_state["password"] == "Expera$$26%==UP"
        ):
            st.session_state["password_correct"] = True
            del st.session_state["password"]  # don't store password
            del st.session_state["username"]
        else:
            st.session_state["password_correct"] = False

    if "password_correct" not in st.session_state:
        # First run, show inputs for username and password.
        st.markdown(f'<div class="main-header" style="margin-top: 50px;">{logo_html}</div>', unsafe_allow_html=True)
        
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            with st.form("Login"):
                st.markdown("### 🔐 Acceso Restringido")
                st.text_input("Usuario", key="username")
                st.text_input("Contraseña", type="password", key="password")
                st.form_submit_button("INGRESAR AL SISTEMA", on_click=password_entered)
                
                if "password_correct" in st.session_state and not st.session_state["password_correct"]:
                    st.error("😕 Usuario o contraseña incorrectos")
            
            st.markdown("<div style='text-align: center; color: #94a3b8; font-size: 0.8rem; margin-top: 20px;'>EXPERA Business Intelligence © 2026</div>", unsafe_allow_html=True)
        return False
    elif not st.session_state["password_correct"]:
        # Password not correct, show input + error.
        st.markdown(f'<div class="main-header" style="margin-top: 50px;">{logo_html}</div>', unsafe_allow_html=True)
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            with st.form("Login"):
                st.markdown("### 🔐 Acceso Restringido")
                st.text_input("Usuario", key="username")
                st.text_input("Contraseña", type="password", key="password")
                st.form_submit_button("INGRESAR AL SISTEMA", on_click=password_entered)
                st.error("😕 Usuario o contraseña incorrectos")
        return False
    else:
        # Password correct.
        return True

# -----------------------------
# Premium UI Layout & Logic
# -----------------------------

# Helper for Logo (must be defined before check_password uses it)
def get_base64_logo(path):
    try:
        with open(path, "rb") as f:
            data = f.read()
            return base64.b64encode(data).decode()
    except:
        return None

logo_base64 = get_base64_logo("EXPERA2.png")
logo_html = f'<img src="data:image/png;base64,{logo_base64}" style="max-height: 150px; width: auto; display: block; margin: 0 auto;">' if logo_base64 else "<h1>EXPERA</h1>"

if check_password():
    # If authenticated, show the dashboard
    
    # Optional: Logout button in sidebar or top
    with st.sidebar:
        st.markdown(f'<div style="text-align: center; padding: 20px;">{logo_html}</div>', unsafe_allow_html=True)
        st.markdown("---")
        st.markdown("### 👤 Sesión Activa")
        st.info("Conectado como Administrador")
        if st.button("Cerrar Sesión"):
            st.session_state["password_correct"] = False
            st.rerun()

    # Header Section with centering optimization
    st.markdown(f"""
    <div class="main-header">
        {logo_html}
    </div>
    """, unsafe_allow_html=True)

    # Search Bar Area - Balanced Layout
    col_spacer_l, col_main, col_spacer_r = st.columns([1, 8, 1])
    
    with col_main:
        c_search, c_filter = st.columns([5, 1.2])
        with c_search:
            search_input = st.text_input(
                label="Buscador Inteligente", 
                placeholder="🔍 ¿Qué buscas hoy? (Marca, Producto, Cliente o Nº)...",
                key="global_search_premium",
                label_visibility="collapsed"
            )
        with c_filter:
            with st.popover("⚡ Filtros"):
                st.markdown("### Configuración de Análisis")
                years_filter = st.slider("Histórico de Años", 1, 15, 5)
                st.info("Estamos procesando toda la base de datos sin límites de registros.")
    
    # Main Dashboard Logic
    if search_input:
        with st.spinner("✨ Procesando inteligencia de datos completa..."):
            df = perform_search(search_input, years_filter)
        
        if not df.empty:
            # Results metrics - New layout
            m1, m2, m3, m4 = st.columns(4)
            
            total_imp = df["Importe Total"].sum()
            total_qty = df["Cant."].sum()
            unique_orders = df["Nº Documento"].nunique()
            avg_price = total_imp / total_qty if total_qty > 0 else 0
    
            m1.metric("Revenue Total", f"S/ {total_imp:,.2f}")
            m2.metric("Volumen Unidades", f"{total_qty:,.0f}")
            m3.metric("Frecuencia Pedidos", f"{unique_orders}")
            m4.metric("Precio Promedio", f"S/ {avg_price:,.2f}")
    
            st.markdown("<br>", unsafe_allow_html=True)
    
            # Tabs with Premium Style
            tab_list, tab_charts = st.tabs(["📋 Auditoría de Ventas", "📈 Insights & Análisis"])
            
            with tab_list:
                # Styled Dataframe
                st.dataframe(
                    df, 
                    use_container_width=True, 
                    height=550,
                    column_config={
                        "Fecha": st.column_config.DateColumn("Fecha", format="DD/MM/YYYY"),
                        "Importe Total": st.column_config.NumberColumn("Importe", format="S/ %.2f"),
                        "Precio Unit.": st.column_config.NumberColumn("Precio", format="S/ %.2f"),
                        "Enlace PDF": st.column_config.LinkColumn("Factura PDF", display_text="Ver Factura"),
                    }
                )
                
                # Action Footer
                st.markdown("<br>", unsafe_allow_html=True)
                f_col1, f_col2, f_col3 = st.columns([1,1,1])
                with f_col2:
                    excel_bytes = to_excel_bytes(df)
                    st.download_button(
                        label="📥 DESCARGAR AUDITORÍA (XLSX)",
                        data=excel_bytes,
                        file_name=f"expera_report_{date.today()}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        width="stretch",
                    )
    
            with tab_charts:
                chart_col1, chart_col2 = st.columns(2)
                
                with chart_col1:
                    st.markdown("#### Distribución de Top Productos")
                    top_p = df.groupby("Producto")["Importe Total"].sum().sort_values(ascending=False).head(10).reset_index()
                    fig_bar = px.bar(
                        top_p, x="Importe Total", y="Producto", orientation='h',
                        color="Importe Total", color_continuous_scale="Viridis",
                        template="plotly_white"
                    )
                    fig_bar.update_layout(showlegend=False, height=450, margin=dict(l=0, r=0, t=10, b=0))
                    st.plotly_chart(fig_bar, use_container_width=True)
                
                with chart_col2:
                    st.markdown("#### Tendencia Temporal (Importe)")
                    # Temporal view
                    df["Fecha"] = pd.to_datetime(df["Fecha"])
                    time_series = df.set_index("Fecha").resample("ME")["Importe Total"].sum().reset_index()
                    fig_line = px.line(
                        time_series, x="Fecha", y="Importe Total",
                        template="plotly_white", line_shape="spline"
                    )
                    fig_line.update_traces(line_color='#3b82f6', line_width=4, fill='tozeroy', fillcolor='rgba(59, 130, 246, 0.1)')
                    fig_line.update_layout(height=450, margin=dict(l=0, r=0, t=10, b=0))
                    st.plotly_chart(fig_line, use_container_width=True)
                    
                # Extra Chart: Top Clients
                st.markdown("---")
                st.markdown("#### Concentración por Clientes")
                top_c = df.groupby("Cliente")["Importe Total"].sum().sort_values(ascending=False).head(15).reset_index()
                fig_pie = px.treemap(top_c, path=['Cliente'], values='Importe Total', color='Importe Total', color_continuous_scale='RdBu')
                st.plotly_chart(fig_pie, use_container_width=True)
                    
        else:
            st.markdown(f"""
            <div style='text-align: center; padding: 100px;'>
                <h2 style='color: #64748b;'>No se encontraron resultados</h2>
                <p style='color: #94a3b8;'>Intenta con otros términos o ajusta el rango de años en filtros.</p>
            </div>
            """, unsafe_allow_html=True)
else:
    st.markdown("<div></div>", unsafe_allow_html=True)
# Implementation complete. Hiding technical logs and SQL as requested.
